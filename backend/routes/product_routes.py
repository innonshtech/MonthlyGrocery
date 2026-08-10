"""
MonthlyGrocery product & pricing routes.

City-aware pricing model:
- Each product carries a `city_prices` dict:
    {
      "Mumbai": {"mrp": 1250, "wholesaler_price": 820, "price": 999, "is_live": True},
      "Pune":   {"mrp": 1200, "wholesaler_price": 810, "price": 979, "is_live": True},
    }
- Consumers scope every request with ?city=<name>. Only products with
  city_prices[<city>].is_live == True are returned, and only that city's mrp/price/discount
  are exposed. `wholesaler_price` is *never* returned to consumers.
- Admins see the full pricing matrix for every city on their `/mine` endpoint.
"""
from datetime import datetime, timezone
from fastapi import APIRouter, Request, HTTPException, Depends, UploadFile, File, Response
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any
from bson import ObjectId
import io
import json

from openpyxl import Workbook, load_workbook

from auth import require_approved_admin
from audit import log_activity

router = APIRouter()


# ---------- Models ----------

class CityPricing(BaseModel):
    mrp: float = 0
    wholesaler_price: float = 0
    price: float = 0
    is_live: bool = True


class ProductCreate(BaseModel):
    name: str
    sku: Optional[str] = ""
    barcode: Optional[str] = ""
    primary_category: str = "Grocery"
    secondary_category: Optional[str] = ""
    brand: Optional[str] = ""
    company: Optional[str] = ""
    description: Optional[str] = ""
    short_description: Optional[str] = ""
    place: Optional[str] = ""
    image_url: Optional[str] = ""
    images: Optional[List[str]] = []
    video_url: Optional[str] = ""
    # City-scoped pricing map (canonical)
    city_prices: Optional[Dict[str, CityPricing]] = Field(default_factory=dict)
    # Legacy defaults — still accepted; new admins should use city_prices
    mrp: float = 0
    wholesaler_price: Optional[float] = 0
    price: float = 0
    gst: Optional[float] = 5
    stock: int = 0
    unit: Optional[str] = "1 Kg"
    quantity_value: Optional[float] = 0
    quantity_unit: Optional[str] = ""
    is_veg: Optional[bool] = True
    featured: Optional[bool] = False
    todays_deal: Optional[bool] = False
    best_seller: Optional[bool] = False
    available: Optional[bool] = True
    search_keywords: Optional[List[str]] = []


# ---------- Helpers ----------

def _compute_discount(mrp: float, price: float) -> int:
    if mrp > 0 and price > 0 and mrp > price:
        return round((1 - price / mrp) * 100)
    return 0


def _city_pricing_out(cp: dict) -> dict:
    """Normalize a city pricing sub-doc for admin responses (includes wholesaler)."""
    if not cp:
        return {}
    mrp = float(cp.get("mrp", 0) or 0)
    price = float(cp.get("price", 0) or 0)
    return {
        "mrp": mrp,
        "wholesaler_price": float(cp.get("wholesaler_price", 0) or 0),
        "price": price,
        "discount_percent": _compute_discount(mrp, price),
        "you_save": round(mrp - price, 2) if mrp > price else 0,
        "is_live": bool(cp.get("is_live", True)),
    }


def _admin_product_out(p: dict) -> dict:
    """Full admin response: exposes wholesaler_price and every city's pricing."""
    out = {k: v for k, v in p.items() if k != "_id"}
    out["id"] = str(p["_id"])
    city_prices = p.get("city_prices", {}) or {}
    out["city_prices"] = {c: _city_pricing_out(cp) for c, cp in city_prices.items()}
    # Legacy fields still present for backward compat
    mrp = float(p.get("mrp", 0) or 0)
    price = float(p.get("price", 0) or 0)
    out["mrp"] = mrp
    out["price"] = price
    out["wholesaler_price"] = float(p.get("wholesaler_price", 0) or 0)
    out["discount_percent"] = _compute_discount(mrp, price)
    out["you_save"] = round(mrp - price, 2) if mrp > price else 0
    return out


def _consumer_product_out(p: dict, city: str) -> Optional[dict]:
    """
    Consumer response. Returns None only when an admin has EXPLICITLY hidden the
    product (`available=False`) or explicitly set the city's price to
    `is_live=False`. Never hides a product just because the current city isn't
    in its `city_prices` map — falls back to the legacy top-level pricing so
    admins don't have to fill every city to make an SKU visible.
    """
    if p.get("available") is False:
        return None
    city_prices = p.get("city_prices", {}) or {}
    cp = city_prices.get(city) if city else None
    if cp is not None:
        # City has explicit pricing — respect the admin's is_live toggle for that city.
        if not cp.get("is_live", True):
            return None
        mrp = float(cp.get("mrp", 0) or 0)
        price = float(cp.get("price", 0) or 0)
    else:
        # No city selected (Pan India) OR no per-city entry for this city
        # → prefer the product's legacy top-level pricing. If that's zero too
        # (common for admins who only fill city rows), fall back to the first
        # live city's pricing so the SKU still shows a real price.
        mrp = float(p.get("mrp", 0) or 0)
        price = float(p.get("price", 0) or 0)
        if price <= 0 and city_prices:
            for _c, _cp in city_prices.items():
                if isinstance(_cp, dict) and _cp.get("is_live", True) and float(_cp.get("price", 0) or 0) > 0:
                    price = float(_cp.get("price", 0) or 0)
                    mrp = float(_cp.get("mrp", 0) or 0) or mrp
                    break

    # Base64 data URLs (from the legacy uploader) bloat every list response by
    # ~500 KB per SKU. Replace them with a lightweight resolver endpoint so the
    # payload stays small and browsers can lazy-load + cache each image.
    raw_img = p.get("image_url", "") or ""
    if raw_img.startswith("data:"):
        img_url = f"/api/products/{str(p['_id'])}/image"
    else:
        img_url = raw_img
    out = {
        "id": str(p["_id"]),
        "name": p.get("name", ""),
        "sku": p.get("sku", ""),
        "primary_category": p.get("primary_category", ""),
        "secondary_category": p.get("secondary_category", ""),
        "brand": p.get("brand", ""),
        "company": p.get("company", ""),
        "short_description": p.get("short_description", ""),
        "place": p.get("place", ""),
        "image_url": img_url,
        "unit": p.get("unit", ""),
        "quantity_value": p.get("quantity_value", 0),
        "quantity_unit": p.get("quantity_unit", ""),
        "gst": p.get("gst", 5),
        "stock": p.get("stock", 0),
        "is_veg": p.get("is_veg", True),
        "featured": p.get("featured", False),
        "todays_deal": p.get("todays_deal", False),
        "best_seller": p.get("best_seller", False),
        "available": p.get("available", True),
        "city": city,
        "mrp": mrp,
        "price": price,
        "discount_percent": _compute_discount(mrp, price),
        "you_save": round(mrp - price, 2) if mrp > price else 0,
    }
    # Heavy detail-page fields — only include when present in the projected doc
    # (i.e. only for /products/{id} detail responses). Keeps /products/all lean.
    for k in ("description", "video_url", "images", "search_keywords"):
        if k in p:
            val = p[k]
            # Also rewrite any base64 images inside the images[] gallery so the
            # detail page benefits from the same lazy-load treatment.
            if k == "images" and isinstance(val, list):
                val = [
                    (f"/api/products/{str(p['_id'])}/image/{i}" if isinstance(u, str) and u.startswith("data:") else u)
                    for i, u in enumerate(val)
                ]
            out[k] = val
    return out


async def _get_platform_shop(db):
    shop = await db.shops.find_one({})
    if not shop:
        raise HTTPException(status_code=500, detail="Storefront not initialized")
    return shop


def _serialize_city_prices(cp: Optional[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    if not cp:
        return {}
    out = {}
    for city, entry in cp.items():
        if isinstance(entry, dict):
            out[str(city).strip()] = {
                "mrp": float(entry.get("mrp", 0) or 0),
                "wholesaler_price": float(entry.get("wholesaler_price", 0) or 0),
                "price": float(entry.get("price", 0) or 0),
                "is_live": bool(entry.get("is_live", True)),
            }
        elif isinstance(entry, CityPricing):
            out[str(city).strip()] = entry.model_dump()
    return out


import base64
from fastapi.responses import RedirectResponse


def _serve_data_url(data_url: str) -> Response:
    """Decode a `data:<mime>;base64,<payload>` string into a Response with raw bytes."""
    try:
        header, payload = data_url.split(",", 1)
        mime = header.split(";")[0].split(":", 1)[1] if ":" in header else "image/png"
        data = base64.b64decode(payload)
    except Exception:
        raise HTTPException(status_code=404, detail="Image not decodable")
    return Response(
        content=data,
        media_type=mime,
        headers={"Cache-Control": "public, max-age=31536000, immutable"},
    )


# ---------- Public: on-demand image resolver ----------

@router.get("/{product_id}/image")
async def resolve_product_image(product_id: str, request: Request):
    """Serve a product's primary image. Handles three storage modes:

    - `image_url` starts with `data:` → decode and stream the bytes (legacy uploads).
    - `image_url` starts with `/api/uploads/` → 302 redirect to the static file.
    - `image_url` is any http(s) URL → 302 redirect.
    - Missing / empty → 404 (frontend falls back to the themed placeholder).
    """
    db = request.app.state.db
    try:
        _id = ObjectId(product_id)
    except Exception:
        raise HTTPException(status_code=404, detail="Product not found")
    p = await db.products.find_one({"_id": _id}, {"image_url": 1})
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    url = (p.get("image_url") or "").strip()
    if not url:
        raise HTTPException(status_code=404, detail="No image")
    if url.startswith("data:"):
        return _serve_data_url(url)
    return RedirectResponse(url=url, status_code=302)


@router.get("/{product_id}/image/{index}")
async def resolve_product_gallery_image(product_id: str, index: int, request: Request):
    """Resolve a specific image from the product's `images[]` gallery — same rules as above."""
    db = request.app.state.db
    try:
        _id = ObjectId(product_id)
    except Exception:
        raise HTTPException(status_code=404, detail="Product not found")
    p = await db.products.find_one({"_id": _id}, {"images": 1})
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    images = p.get("images") or []
    if index < 0 or index >= len(images):
        raise HTTPException(status_code=404, detail="No such image")
    url = (images[index] or "").strip()
    if not url:
        raise HTTPException(status_code=404, detail="No image")
    if url.startswith("data:"):
        return _serve_data_url(url)
    return RedirectResponse(url=url, status_code=302)


# ---------- Admin CRUD ----------

@router.get("/mine")
async def list_my_products(request: Request, user: dict = Depends(require_approved_admin())):
    db = request.app.state.db
    shop = await _get_platform_shop(db)
    products = await db.products.find({"shop_id": str(shop["_id"])}).sort("created_at", -1).to_list(2000)
    return {"products": [_admin_product_out(p) for p in products]}


@router.get("/mine/visibility")
async def visibility_diagnostic(request: Request, user: dict = Depends(require_approved_admin())):
    """Explain how many SKUs are visible to consumers vs hidden and why.
    Helps admins debug 'my SKUs aren't showing on the shop' quickly."""
    db = request.app.state.db
    shop = await _get_platform_shop(db)
    total = await db.products.count_documents({"shop_id": str(shop["_id"])})
    unavailable = await db.products.count_documents({
        "shop_id": str(shop["_id"]),
        "available": {"$ne": True},
    })
    no_price = await db.products.count_documents({
        "shop_id": str(shop["_id"]),
        "available": True,
        "$and": [
            {"$or": [{"price": {"$lte": 0}}, {"price": {"$exists": False}}]},
            {"$or": [{"city_prices": {}}, {"city_prices": {"$exists": False}}]},
        ],
    })
    all_city_hidden = await db.products.count_documents({
        "shop_id": str(shop["_id"]),
        "available": True,
        "city_prices.Mumbai.is_live": {"$ne": True},
        "city_prices.Pune.is_live": {"$ne": True},
        "city_prices.Bengaluru.is_live": {"$ne": True},
        "$or": [{"price": {"$lte": 0}}, {"price": {"$exists": False}}],
    })
    visible = total - unavailable - no_price
    return {
        "total": total,
        "visible_to_consumers": max(0, visible),
        "hidden": {
            "explicitly_unavailable": unavailable,
            "no_price_set": no_price,
            "all_cities_paused": all_city_hidden,
        },
    }


@router.post("/mine/publish-all")
async def publish_all(request: Request, user: dict = Depends(require_approved_admin())):
    """Force every SKU in the catalog to be visible to consumers.

    Sets `available=True`, marks every `city_prices[*].is_live=True`, and — for
    SKUs whose legacy top-level price is 0 — backfills it from the first live
    city entry so the /shop grid always shows a real ₹ figure.
    """
    db = request.app.state.db
    shop = await _get_platform_shop(db)

    # Simple field: mark every SKU available. One update, atomic.
    r1 = await db.products.update_many(
        {"shop_id": str(shop["_id"]), "available": {"$ne": True}},
        {"$set": {"available": True}},
    )

    # Complex fields: flip is_live for every city + backfill legacy price where missing.
    fixed_prices = 0
    unpaused_cities = 0
    async for p in db.products.find({"shop_id": str(shop["_id"])}):
        updates: Dict[str, Any] = {}
        city_prices = p.get("city_prices", {}) or {}
        changed_cp = False
        first_price = None
        first_mrp = None
        first_wholesale = None
        for cname, cp in list(city_prices.items()):
            if not isinstance(cp, dict):
                continue
            if not cp.get("is_live", True):
                cp["is_live"] = True
                changed_cp = True
                unpaused_cities += 1
            if first_price is None and float(cp.get("price", 0) or 0) > 0:
                first_price = float(cp.get("price", 0) or 0)
                first_mrp = float(cp.get("mrp", 0) or 0)
                first_wholesale = float(cp.get("wholesaler_price", 0) or 0)
        if changed_cp:
            updates["city_prices"] = city_prices
        if float(p.get("price", 0) or 0) <= 0 and first_price:
            updates["price"] = first_price
            updates["mrp"] = first_mrp or first_price
            if first_wholesale:
                updates["wholesaler_price"] = first_wholesale
            fixed_prices += 1
        if updates:
            await db.products.update_one({"_id": p["_id"]}, {"$set": updates})

    await log_activity(
        db, user, action="sku.publish_all", resource_type="product",
        resource_id="-", resource_name=f"{r1.modified_count} SKUs unhidden",
        metadata={
            "made_available": r1.modified_count,
            "unpaused_cities": unpaused_cities,
            "backfilled_prices": fixed_prices,
        },
    )
    return {
        "success": True,
        "made_available": r1.modified_count,
        "unpaused_cities": unpaused_cities,
        "backfilled_prices": fixed_prices,
    }


@router.post("/mine")
async def create_product(payload: ProductCreate, request: Request, user: dict = Depends(require_approved_admin())):
    db = request.app.state.db
    shop = await _get_platform_shop(db)
    data = payload.model_dump()
    data["city_prices"] = _serialize_city_prices(data.get("city_prices"))
    data["shop_id"] = str(shop["_id"])
    data["shop_name"] = shop.get("shop_name", "MonthlyGrocery")
    data["created_at"] = datetime.now(timezone.utc).isoformat()
    data["created_by"] = user["id"]
    result = await db.products.insert_one(data)
    doc = await db.products.find_one({"_id": result.inserted_id})
    await log_activity(db, user, action="sku.create", resource_type="product",
                       resource_id=str(result.inserted_id), resource_name=data.get("name", ""))
    return {"product": _admin_product_out(doc)}


@router.put("/mine/{product_id}")
async def update_product(product_id: str, payload: ProductCreate, request: Request, user: dict = Depends(require_approved_admin())):
    db = request.app.state.db
    try:
        _id = ObjectId(product_id)
    except Exception:
        raise HTTPException(status_code=404, detail="SKU not found")
    prod = await db.products.find_one({"_id": _id})
    if not prod:
        raise HTTPException(status_code=404, detail="SKU not found")
    data = payload.model_dump()
    data["city_prices"] = _serialize_city_prices(data.get("city_prices"))
    await db.products.update_one({"_id": _id}, {"$set": data})
    doc = await db.products.find_one({"_id": _id})
    await log_activity(db, user, action="sku.update", resource_type="product",
                       resource_id=product_id, resource_name=data.get("name", ""))
    return {"product": _admin_product_out(doc)}


@router.delete("/mine/{product_id}")
async def delete_product(product_id: str, request: Request, user: dict = Depends(require_approved_admin())):
    db = request.app.state.db
    try:
        _id = ObjectId(product_id)
    except Exception:
        raise HTTPException(status_code=404, detail="SKU not found")
    prod = await db.products.find_one({"_id": _id})
    result = await db.products.delete_one({"_id": _id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="SKU not found")
    await log_activity(db, user, action="sku.delete", resource_type="product",
                       resource_id=product_id, resource_name=(prod or {}).get("name", ""))
    return {"success": True}


class BulkDeleteRequest(BaseModel):
    # Cap ids array to protect against pathological payloads.
    ids: Optional[List[str]] = Field(default=None, max_length=1000)
    all: bool = False
    confirm: Optional[str] = None


@router.post("/mine/bulk-delete")
async def bulk_delete_products(payload: BulkDeleteRequest, request: Request, user: dict = Depends(require_approved_admin())):
    """
    Delete many SKUs in one call.
    - Pass `ids=[...]` to remove a specific selection.
    - Pass `all=True, confirm='DELETE ALL'` to wipe the whole catalog.
    """
    db = request.app.state.db
    shop = await _get_platform_shop(db)

    if payload.all:
        if (payload.confirm or "").strip().upper() != "DELETE ALL":
            raise HTTPException(status_code=400, detail="To delete every SKU, pass confirm='DELETE ALL'.")
        total = await db.products.count_documents({"shop_id": str(shop["_id"])})
        result = await db.products.delete_many({"shop_id": str(shop["_id"])})
        await log_activity(
            db, user, action="sku.bulk_delete_all", resource_type="product",
            resource_id="-", resource_name=f"{result.deleted_count}/{total} SKUs",
            metadata={"deleted": result.deleted_count, "total": total},
        )
        return {"success": True, "deleted": result.deleted_count, "mode": "all"}

    ids = [i for i in (payload.ids or []) if i]
    if not ids:
        raise HTTPException(status_code=400, detail="Provide `ids` to delete, or `all=true` to wipe everything.")
    try:
        obj_ids = [ObjectId(i) for i in ids]
    except Exception:
        raise HTTPException(status_code=400, detail="One or more ids are invalid.")
    result = await db.products.delete_many({
        "_id": {"$in": obj_ids},
        "shop_id": str(shop["_id"]),
    })
    await log_activity(
        db, user, action="sku.bulk_delete", resource_type="product",
        resource_id="-", resource_name=f"{result.deleted_count} SKUs",
        metadata={"requested": len(ids), "deleted": result.deleted_count},
    )
    return {"success": True, "deleted": result.deleted_count, "mode": "ids"}


# ---------- Search / consumer listing ----------

def _build_search_regex(q: str):
    import re as _re
    tokens = [t for t in _re.split(r"\s+", q.strip()) if t]
    if not tokens:
        return None
    ors = []
    for t in tokens:
        pat = _re.escape(t)
        for field in ["name", "brand", "company", "primary_category", "secondary_category",
                      "description", "short_description", "place", "search_keywords"]:
            ors.append({field: {"$regex": pat, "$options": "i"}})
    return {"$or": ors}


@router.get("/cities")
async def list_live_cities(request: Request):
    """Every city that has at least one live SKU. Consumer city-picker source of truth."""
    db = request.app.state.db
    pipeline = [
        {"$match": {"available": True}},
        {"$project": {"cp": {"$objectToArray": {"$ifNull": ["$city_prices", {}]}}}},
        {"$unwind": "$cp"},
        {"$match": {"cp.v.is_live": True}},
        {"$group": {"_id": "$cp.k", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    docs = await db.products.aggregate(pipeline).to_list(200)
    cities = [{"name": d["_id"], "sku_count": d["count"]} for d in docs if d["_id"]]
    return {"cities": cities}


@router.get("/all")
async def list_all_products(
    request: Request,
    city: Optional[str] = None,
    category: Optional[str] = None,
    secondary: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 500,
):
    db = request.app.state.db
    base = {"available": True}
    if category:
        base["primary_category"] = category
    if secondary:
        base["secondary_category"] = secondary
    if q:
        r = _build_search_regex(q)
        if r:
            base = {"$and": [base, r]}

    # City filtering — only products live in this city. If no city is passed, return everything.
    if city:
        city_field = f"city_prices.{city}.is_live"
        base = {"$and": [base, {"$or": [
            {city_field: True},
            {"city_prices": {"$in": [None, {}]}},  # legacy products stay visible
        ]}]}

    # Projection — omit heavy fields (description, video_url, images[], search_keywords[])
    # from the list endpoint to keep the guest-facing /shop response small & fast.
    # These fields are still returned by GET /products/{id} for the detail page.
    projection = {
        "name": 1, "sku": 1, "primary_category": 1, "secondary_category": 1,
        "brand": 1, "company": 1, "short_description": 1, "place": 1,
        "image_url": 1, "unit": 1, "quantity_value": 1, "quantity_unit": 1,
        "gst": 1, "stock": 1, "is_veg": 1, "featured": 1, "todays_deal": 1,
        "best_seller": 1, "available": 1, "mrp": 1, "price": 1, "city_prices": 1,
    }
    products = await db.products.find(base, projection).limit(limit).to_list(limit)
    out = []
    for p in products:
        rendered = _consumer_product_out(p, city or "")
        if rendered:
            out.append(rendered)
    return {"products": out}


@router.get("/tree")
async def category_tree(request: Request, city: Optional[str] = None):
    """Primary → secondary category hierarchy, filtered by city if provided."""
    db = request.app.state.db
    match = {"available": True}
    if city:
        city_field = f"city_prices.{city}.is_live"
        match = {"$and": [match, {"$or": [
            {city_field: True},
            {"city_prices": {"$in": [None, {}]}},
        ]}]}
    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": {"primary": "$primary_category", "secondary": "$secondary_category"},
            "count": {"$sum": 1},
        }},
    ]
    docs = await db.products.aggregate(pipeline).to_list(500)
    tree = {}
    for d in docs:
        primary = d["_id"].get("primary") or "General"
        secondary = d["_id"].get("secondary") or ""
        if primary not in tree:
            tree[primary] = {"name": primary, "count": 0, "subs": {}}
        tree[primary]["count"] += d["count"]
        if secondary:
            tree[primary]["subs"][secondary] = tree[primary]["subs"].get(secondary, 0) + d["count"]
    result = []
    for primary, node in sorted(tree.items(), key=lambda kv: -kv[1]["count"]):
        subs = [{"name": s, "count": c} for s, c in sorted(node["subs"].items(), key=lambda kv: -kv[1])]
        result.append({"name": primary, "count": node["count"], "subs": subs})
    return {"categories": result}


# ---------- Excel export & import ----------

EXCEL_COLUMNS = [
    "product_name", "city", "mrp", "purchase_price", "selling_price", "is_live",
    "primary_category", "secondary_category", "brand", "company", "unit",
    "quantity_value", "quantity_unit", "stock", "gst", "description", "short_description",
    "image_url", "video_url", "place", "search_keywords",
    "featured", "todays_deal", "best_seller", "available",
]


@router.get("/export-excel")
async def export_excel(request: Request, user: dict = Depends(require_approved_admin())):
    """Export the entire catalog as an Excel file (one row per product × city)."""
    db = request.app.state.db
    shop = await _get_platform_shop(db)
    products = await db.products.find({"shop_id": str(shop["_id"])}).sort("created_at", -1).to_list(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "MonthlyGrocery SKUs"
    ws.append(EXCEL_COLUMNS)

    for p in products:
        city_prices = p.get("city_prices", {}) or {}
        rows_written = 0
        for city_name, cp in city_prices.items():
            ws.append([
                p.get("name", ""), city_name,
                float(cp.get("mrp", 0) or 0),
                float(cp.get("wholesaler_price", 0) or 0),
                float(cp.get("price", 0) or 0),
                "yes" if cp.get("is_live", True) else "no",
                p.get("primary_category", ""), p.get("secondary_category", ""),
                p.get("brand", ""), p.get("company", ""), p.get("unit", ""),
                float(p.get("quantity_value", 0) or 0), p.get("quantity_unit", ""),
                int(p.get("stock", 0) or 0), float(p.get("gst", 5) or 5),
                p.get("description", ""), p.get("short_description", ""),
                p.get("image_url", ""), p.get("video_url", ""), p.get("place", ""),
                ", ".join(p.get("search_keywords", []) or []),
                "yes" if p.get("featured") else "no",
                "yes" if p.get("todays_deal") else "no",
                "yes" if p.get("best_seller") else "no",
                "yes" if p.get("available", True) else "no",
            ])
            rows_written += 1
        # If product has no city_prices yet, still export a row using legacy fields
        if rows_written == 0:
            ws.append([
                p.get("name", ""), "",
                float(p.get("mrp", 0) or 0),
                float(p.get("wholesaler_price", 0) or 0),
                float(p.get("price", 0) or 0),
                "no",
                p.get("primary_category", ""), p.get("secondary_category", ""),
                p.get("brand", ""), p.get("company", ""), p.get("unit", ""),
                float(p.get("quantity_value", 0) or 0), p.get("quantity_unit", ""),
                int(p.get("stock", 0) or 0), float(p.get("gst", 5) or 5),
                p.get("description", ""), p.get("short_description", ""),
                p.get("image_url", ""), p.get("video_url", ""), p.get("place", ""),
                ", ".join(p.get("search_keywords", []) or []),
                "yes" if p.get("featured") else "no",
                "yes" if p.get("todays_deal") else "no",
                "yes" if p.get("best_seller") else "no",
                "yes" if p.get("available", True) else "no",
            ])

    # Column widths
    widths = {"A": 34, "B": 14, "C": 10, "D": 14, "E": 12, "F": 8, "G": 20, "H": 20, "I": 16, "J": 16,
              "K": 12, "L": 12, "M": 12, "N": 8, "O": 6, "P": 40, "Q": 30, "R": 40, "S": 30, "T": 14,
              "U": 40, "V": 10, "W": 12, "X": 10, "Y": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"monthlygrocery-catalog-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/excel-template")
async def excel_template(request: Request, user: dict = Depends(require_approved_admin())):
    """Empty template with headers + 2 sample rows to guide the admin."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MonthlyGrocery SKUs"
    ws.append(EXCEL_COLUMNS)
    ws.append([
        "Aashirvaad Shudh Chakki Atta 10kg", "Mumbai", 595, 380, 479, "yes",
        "Atta & Rice", "Atta", "Aashirvaad", "ITC", "10 Kg",
        10, "kg", 100, 5, "100% chakki whole wheat atta.", "Whole wheat atta",
        "", "", "Bhopal, MP", "atta, आटा, gehu, wheat flour",
        "yes", "no", "yes", "yes",
    ])
    ws.append([
        "Aashirvaad Shudh Chakki Atta 10kg", "Pune", 595, 380, 469, "yes",
        "Atta & Rice", "Atta", "Aashirvaad", "ITC", "10 Kg",
        10, "kg", 100, 5, "100% chakki whole wheat atta.", "Whole wheat atta",
        "", "", "Bhopal, MP", "atta, आटा, gehu, wheat flour",
        "yes", "no", "yes", "yes",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="monthlygrocery-template.xlsx"'},
    )


def _yes(v) -> bool:
    if isinstance(v, bool):
        return v
    s = str(v or "").strip().lower()
    return s in ("yes", "y", "true", "1", "live")


def _yes_default_true(v) -> bool:
    """Same as _yes but treats missing / blank / None as True.
    Used for import defaults (`available`, `is_live`) so an admin who leaves the
    cell blank ends up with a VISIBLE SKU instead of an invisible one."""
    if v is None:
        return True
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s == "":
        return True
    return s in ("yes", "y", "true", "1", "live")


@router.post("/import-excel")
async def import_excel(request: Request, file: UploadFile = File(...), user: dict = Depends(require_approved_admin())):
    """
    Bulk-import SKUs. One row per (product × city).
    - Matches existing products by lowercased name.
    - Updates that product's city_prices[<city>].
    - Non-price fields on the row overwrite the product's non-price fields
      (the LAST row for a given product wins for those).
    """
    db = request.app.state.db
    shop = await _get_platform_shop(db)

    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if "product_name" not in header:
        raise HTTPException(status_code=400, detail="Missing required column 'product_name'.")

    def cell(row, key):
        try:
            idx = header.index(key)
        except ValueError:
            return None
        return row[idx].value if idx < len(row) else None

    created = 0
    updated = 0
    rows_processed = 0
    errors = []

    for row in ws.iter_rows(min_row=2):
        rows_processed += 1
        name = str(cell(row, "product_name") or "").strip()
        if not name:
            continue
        try:
            city_name = str(cell(row, "city") or "").strip()
            mrp = float(cell(row, "mrp") or 0)
            purchase = float(cell(row, "purchase_price") or 0)
            selling = float(cell(row, "selling_price") or 0)
            is_live = _yes_default_true(cell(row, "is_live"))

            kw_raw = str(cell(row, "search_keywords") or "").strip()
            keywords = [k.strip() for k in kw_raw.replace(";", ",").split(",") if k.strip()]

            base_fields = {
                "name": name,
                "primary_category": str(cell(row, "primary_category") or "Grocery"),
                "secondary_category": str(cell(row, "secondary_category") or ""),
                "brand": str(cell(row, "brand") or ""),
                "company": str(cell(row, "company") or ""),
                "unit": str(cell(row, "unit") or ""),
                "quantity_value": float(cell(row, "quantity_value") or 0),
                "quantity_unit": str(cell(row, "quantity_unit") or ""),
                "stock": int(float(cell(row, "stock") or 0)),
                "gst": float(cell(row, "gst") or 5),
                "description": str(cell(row, "description") or ""),
                "short_description": str(cell(row, "short_description") or ""),
                "image_url": str(cell(row, "image_url") or ""),
                "video_url": str(cell(row, "video_url") or ""),
                "place": str(cell(row, "place") or ""),
                "search_keywords": keywords,
                "featured": _yes(cell(row, "featured")),
                "todays_deal": _yes(cell(row, "todays_deal")),
                "best_seller": _yes(cell(row, "best_seller")),
                "available": _yes_default_true(cell(row, "available")),
            }

            existing = await db.products.find_one({
                "shop_id": str(shop["_id"]),
                "name": {"$regex": f"^{name}$", "$options": "i"},
            })

            city_entry = {
                "mrp": mrp,
                "wholesaler_price": purchase,
                "price": selling,
                "is_live": is_live,
            }

            if existing:
                cp = dict(existing.get("city_prices", {}) or {})
                if city_name:
                    cp[city_name] = city_entry
                update = {**base_fields, "city_prices": cp}
                # keep legacy defaults in sync (use first city's price)
                if city_name and is_live:
                    update["mrp"] = mrp
                    update["price"] = selling
                    update["wholesaler_price"] = purchase
                await db.products.update_one({"_id": existing["_id"]}, {"$set": update})
                updated += 1
            else:
                doc = {
                    **base_fields,
                    "shop_id": str(shop["_id"]),
                    "shop_name": shop.get("shop_name", "MonthlyGrocery"),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": user["id"],
                    "city_prices": {city_name: city_entry} if city_name else {},
                    "mrp": mrp, "price": selling, "wholesaler_price": purchase,
                    "images": [], "is_veg": True,
                }
                await db.products.insert_one(doc)
                created += 1
        except Exception as ex:
            errors.append({"row": rows_processed + 1, "error": str(ex)})

    await _log_bulk_import(db, user, rows_processed, created, updated)
    return {
        "success": True,
        "rows_processed": rows_processed,
        "created": created,
        "updated": updated,
        "errors": errors[:20],
    }


async def _log_bulk_import(db, user, rows_processed, created, updated):
    await log_activity(db, user, action="sku.bulk_import", resource_type="product",
                       resource_name=f"{created} created / {updated} updated",
                       metadata={"rows_processed": rows_processed, "created": created, "updated": updated})


# ---------- Public single-product ----------

@router.get("/{product_id}")
async def get_product(product_id: str, request: Request, city: Optional[str] = None):
    db = request.app.state.db
    try:
        p = await db.products.find_one({"_id": ObjectId(product_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="SKU not found")
    if not p:
        raise HTTPException(status_code=404, detail="SKU not found")
    rendered = _consumer_product_out(p, city or "")
    if not rendered:
        raise HTTPException(status_code=404, detail="Not available in this city")
    return {"product": rendered}
