"""
Iteration 6 — city-aware pricing + Excel import/export tests.
Uses same JWT-mint pattern as backend_test.py to avoid Twilio.
"""
import os
import io
import time
import jwt as pyjwt
import pytest
import requests
from datetime import datetime, timezone, timedelta
from pymongo import MongoClient
from bson import ObjectId
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv("/app/backend/.env")

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/") if os.environ.get("REACT_APP_BACKEND_URL") else "http://localhost:8001"
# Prefer frontend/.env's REACT_APP_BACKEND_URL if set
try:
    load_dotenv("/app/frontend/.env")
    if os.environ.get("REACT_APP_BACKEND_URL"):
        BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
except Exception:
    pass

API = f"{BASE_URL}/api"

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ["JWT_SECRET"]
SUPER_MOBILE = os.environ.get("SUPER_ADMIN_MOBILE", "+919833833498")

_mongo = MongoClient(MONGO_URL)
db = _mongo[DB_NAME]


def _mint_token(user_id, mobile, role, status="approved"):
    return pyjwt.encode({
        "sub": user_id, "mobile": mobile, "role": role, "status": status,
        "type": "access",
        "exp": datetime.now(timezone.utc) + timedelta(hours=1),
    }, JWT_SECRET, algorithm="HS256")


def _sess(uid, mobile, role, status="approved"):
    tok = _mint_token(uid, mobile, role, status)
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


def _mk_user(role, status="approved", name="Test"):
    tail = str(int(time.time() * 1000) % 100000000).zfill(8)
    mobile = "+9198" + tail
    doc = {"mobile": mobile, "name": f"TEST_{name}", "role": role, "status": status,
           "address": "", "city": "", "pincode": "",
           "created_at": datetime.now(timezone.utc).isoformat()}
    res = db.users.insert_one(doc)
    return {"id": str(res.inserted_id), "mobile": mobile, "role": role, "status": status}


@pytest.fixture(scope="module")
def consumer():
    u = _mk_user("consumer", "approved", "Consumer")
    yield u
    db.users.delete_one({"_id": ObjectId(u["id"])})
    db.carts.delete_many({"user_id": u["id"]})


@pytest.fixture(scope="module")
def admin_approved():
    u = _mk_user("admin", "approved", "AdminApproved")
    yield u
    db.users.delete_one({"_id": ObjectId(u["id"])})


@pytest.fixture(scope="module")
def super_admin():
    sa = db.users.find_one({"mobile": SUPER_MOBILE})
    assert sa
    return {"id": str(sa["_id"]), "mobile": sa["mobile"], "role": "super_admin", "status": "approved"}


@pytest.fixture
def cs(consumer):
    return _sess(consumer["id"], consumer["mobile"], "consumer")


@pytest.fixture
def aas(admin_approved):
    return _sess(admin_approved["id"], admin_approved["mobile"], "admin")


@pytest.fixture
def sas(super_admin):
    return _sess(super_admin["id"], super_admin["mobile"], "super_admin")


# ---------- Cities endpoint ----------
class TestCitiesEndpoint:
    def test_cities_public_no_auth(self):
        r = requests.get(f"{API}/products/cities", timeout=15)
        assert r.status_code == 200, r.text
        data = r.json()
        assert "cities" in data
        assert isinstance(data["cities"], list)

    def test_cities_returns_seeded_three(self):
        r = requests.get(f"{API}/products/cities", timeout=15)
        cities = r.json()["cities"]
        names = {c["name"] for c in cities}
        assert {"Mumbai", "Pune", "Bengaluru"}.issubset(names), f"got {names}"
        for c in cities:
            assert "name" in c and "sku_count" in c
            assert isinstance(c["sku_count"], int)
            assert c["sku_count"] > 0


# ---------- /all with city filter ----------
class TestConsumerListingByCity:
    def test_all_mumbai_hides_wholesaler(self):
        r = requests.get(f"{API}/products/all", params={"city": "Mumbai"}, timeout=15)
        assert r.status_code == 200, r.text
        prods = r.json()["products"]
        assert len(prods) > 0
        for p in prods:
            assert "wholesaler_price" not in p, f"leaked wholesaler_price in {p['name']}"
            assert "city_prices" not in p, f"leaked city_prices in {p['name']}"
            assert p["city"] == "Mumbai"
            for k in ("mrp", "price", "discount_percent"):
                assert k in p

    def test_all_nonexistent_city_empty(self):
        r = requests.get(f"{API}/products/all", params={"city": "AtlantisXYZ"}, timeout=15)
        assert r.status_code == 200
        assert r.json()["products"] == []

    def test_all_no_city_backwards_compat(self):
        r = requests.get(f"{API}/products/all", timeout=15)
        assert r.status_code == 200
        prods = r.json()["products"]
        assert len(prods) > 0
        for p in prods:
            assert "wholesaler_price" not in p
            for k in ("mrp", "price"):
                assert k in p

    def test_get_single_product_hides_wholesaler(self):
        prods = requests.get(f"{API}/products/all", params={"city": "Mumbai"}, timeout=15).json()["products"]
        pid = prods[0]["id"]
        r = requests.get(f"{API}/products/{pid}", params={"city": "Mumbai"}, timeout=15)
        assert r.status_code == 200
        p = r.json()["product"]
        assert "wholesaler_price" not in p
        assert "city_prices" not in p
        assert p["city"] == "Mumbai"


# ---------- Admin /mine full pricing matrix ----------
class TestAdminMineFullMatrix:
    def test_mine_exposes_city_prices_with_wholesaler(self, aas):
        r = aas.get(f"{API}/products/mine", timeout=15)
        assert r.status_code == 200
        prods = r.json()["products"]
        assert len(prods) > 0
        # Find a product that has city_prices
        has_cp = [p for p in prods if p.get("city_prices")]
        assert has_cp, "no product exposes city_prices to admin"
        p = has_cp[0]
        for city_name, cp in p["city_prices"].items():
            for k in ("mrp", "wholesaler_price", "price", "is_live", "discount_percent"):
                assert k in cp, f"missing {k} in city_prices[{city_name}]"


# ---------- Toggle is_live via PUT ----------
class TestToggleCityLive:
    def test_toggle_is_live_hides_and_restores(self, aas):
        # Create a fresh product live in Mumbai only
        payload = {
            "name": f"TEST_CityToggle_{int(time.time()*1000)}",
            "primary_category": "Grocery",
            "city_prices": {
                "Mumbai": {"mrp": 100, "wholesaler_price": 50, "price": 80, "is_live": True},
            },
            "mrp": 100, "price": 80, "wholesaler_price": 50,
        }
        r = aas.post(f"{API}/products/mine", json=payload, timeout=15)
        assert r.status_code == 200, r.text
        pid = r.json()["product"]["id"]
        try:
            # Visible in Mumbai
            all_mumbai = requests.get(f"{API}/products/all", params={"city": "Mumbai"}, timeout=15).json()["products"]
            assert any(p["id"] == pid for p in all_mumbai), "created product not in Mumbai listing"

            # Toggle is_live=false
            update = dict(payload)
            update["city_prices"]["Mumbai"]["is_live"] = False
            r2 = aas.put(f"{API}/products/mine/{pid}", json=update, timeout=15)
            assert r2.status_code == 200, r2.text

            all_mumbai2 = requests.get(f"{API}/products/all", params={"city": "Mumbai"}, timeout=15).json()["products"]
            assert not any(p["id"] == pid for p in all_mumbai2), "product still visible after is_live=false"

            # Single product endpoint returns 404
            r3 = requests.get(f"{API}/products/{pid}", params={"city": "Mumbai"}, timeout=15)
            assert r3.status_code == 404

            # Restore
            update["city_prices"]["Mumbai"]["is_live"] = True
            r4 = aas.put(f"{API}/products/mine/{pid}", json=update, timeout=15)
            assert r4.status_code == 200
            all_mumbai3 = requests.get(f"{API}/products/all", params={"city": "Mumbai"}, timeout=15).json()["products"]
            assert any(p["id"] == pid for p in all_mumbai3)
        finally:
            aas.delete(f"{API}/products/mine/{pid}", timeout=15)

    def test_cities_recompute_drops_city_with_no_live_skus(self, aas):
        # Create a product live only in a novel city, then toggle off — city should drop out
        unique_city = f"TestCity{int(time.time()*1000)}"
        payload = {
            "name": f"TEST_NovelCity_{int(time.time()*1000)}",
            "primary_category": "Grocery",
            "city_prices": {unique_city: {"mrp": 100, "wholesaler_price": 50, "price": 80, "is_live": True}},
        }
        r = aas.post(f"{API}/products/mine", json=payload, timeout=15)
        pid = r.json()["product"]["id"]
        try:
            cities = requests.get(f"{API}/products/cities", timeout=15).json()["cities"]
            assert unique_city in {c["name"] for c in cities}

            # Turn off live
            payload["city_prices"][unique_city]["is_live"] = False
            aas.put(f"{API}/products/mine/{pid}", json=payload, timeout=15)
            cities2 = requests.get(f"{API}/products/cities", timeout=15).json()["cities"]
            assert unique_city not in {c["name"] for c in cities2}, "city with no live SKUs still listed"
        finally:
            aas.delete(f"{API}/products/mine/{pid}", timeout=15)


# ---------- Excel export ----------
class TestExcelExport:
    def test_export_super_admin_ok(self, sas):
        r = sas.get(f"{API}/products/export-excel", timeout=30)
        assert r.status_code == 200, r.text
        assert "spreadsheet" in r.headers.get("content-type", "")
        assert len(r.content) > 500
        # Parseable
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "product_name" in header
        assert "city" in header
        assert "purchase_price" in header

    def test_export_admin_ok(self, aas):
        r = aas.get(f"{API}/products/export-excel", timeout=30)
        assert r.status_code == 200

    def test_export_consumer_forbidden(self, cs):
        r = cs.get(f"{API}/products/export-excel", timeout=15)
        assert r.status_code == 403


# ---------- Excel template ----------
class TestExcelTemplate:
    def test_template_headers_and_samples(self, aas):
        r = aas.get(f"{API}/products/excel-template", timeout=15)
        assert r.status_code == 200
        assert "spreadsheet" in r.headers.get("content-type", "")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 3  # header + 2 samples
        assert rows[0][0] == "product_name"
        assert rows[1][0] and rows[2][0]  # sample rows have data


# ---------- Excel import ----------
class TestExcelImport:
    def _make_xlsx(self, rows):
        wb = Workbook()
        ws = wb.active
        headers = ["product_name", "city", "mrp", "purchase_price", "selling_price", "is_live",
                   "primary_category", "secondary_category", "brand", "company"]
        ws.append(headers)
        for r in rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def test_import_creates_and_updates(self, aas):
        unique_name = f"TEST_ImportSKU_{int(time.time()*1000)}"
        # Create via import
        content = self._make_xlsx([
            [unique_name, "Mumbai", 200, 120, 175, "yes", "Grocery", "Test", "TestBrand", "TestCo"],
        ])
        r = aas.post(
            f"{API}/products/import-excel",
            files={"file": ("t.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30,
        )
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["rows_processed"] == 1
        assert d["created"] == 1
        assert d["updated"] == 0

        # Now update same product (case-insensitive match)
        content2 = self._make_xlsx([
            [unique_name.lower(), "Mumbai", 210, 125, 180, "yes", "Grocery", "Test", "TestBrand", "TestCo"],
            [unique_name, "Pune", 210, 125, 185, "yes", "Grocery", "Test", "TestBrand", "TestCo"],
        ])
        r2 = aas.post(
            f"{API}/products/import-excel",
            files={"file": ("t.xlsx", content2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30,
        )
        assert r2.status_code == 200
        d2 = r2.json()
        assert d2["rows_processed"] == 2
        assert d2["updated"] == 2  # both rows update same existing product
        assert d2["created"] == 0

        # Cleanup
        db.products.delete_many({"name": {"$regex": f"^{unique_name}$", "$options": "i"}})

    def test_import_consumer_forbidden(self, cs):
        content = self._make_xlsx([["X", "Mumbai", 1, 1, 1, "yes", "", "", "", ""]])
        r = cs.post(
            f"{API}/products/import-excel",
            files={"file": ("t.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=15,
        )
        assert r.status_code == 403


# ---------- Cart with city ----------
class TestCartCityAware:
    def test_cart_add_persists_city(self, cs):
        cs.post(f"{API}/cart/clear", timeout=15)
        prods = requests.get(f"{API}/products/all", params={"city": "Mumbai"}, timeout=15).json()["products"]
        assert prods
        pid = prods[0]["id"]
        r = cs.post(f"{API}/cart/add", json={"product_id": pid, "quantity": 1, "city": "Mumbai"}, timeout=15)
        assert r.status_code == 200
        d = r.json()
        assert d["city"] == "Mumbai"
        assert d["items"][0]["price"] == prods[0]["price"]
        cs.post(f"{API}/cart/clear", timeout=15)

    def test_changing_city_clears_items(self, cs):
        cs.post(f"{API}/cart/clear", timeout=15)
        mumbai_prods = requests.get(f"{API}/products/all", params={"city": "Mumbai"}, timeout=15).json()["products"]
        pid = mumbai_prods[0]["id"]
        cs.post(f"{API}/cart/add", json={"product_id": pid, "quantity": 1, "city": "Mumbai"}, timeout=15)

        # Switch city → items should be wiped
        r = cs.post(f"{API}/cart/set-city", json={"city": "Pune"}, timeout=15)
        assert r.status_code == 200
        d = r.json()
        assert d["city"] == "Pune"
        assert d["items"] == []
        cs.post(f"{API}/cart/clear", timeout=15)

    def test_cart_uses_city_price(self, cs, aas):
        # Create product with different Mumbai vs Pune prices
        payload = {
            "name": f"TEST_CartCityPrice_{int(time.time()*1000)}",
            "primary_category": "Grocery",
            "city_prices": {
                "Mumbai": {"mrp": 200, "wholesaler_price": 100, "price": 150, "is_live": True},
                "Pune": {"mrp": 200, "wholesaler_price": 100, "price": 175, "is_live": True},
            },
            "mrp": 200, "price": 150, "wholesaler_price": 100,
        }
        r = aas.post(f"{API}/products/mine", json=payload, timeout=15)
        pid = r.json()["product"]["id"]
        try:
            cs.post(f"{API}/cart/clear", timeout=15)
            r1 = cs.post(f"{API}/cart/add", json={"product_id": pid, "quantity": 1, "city": "Mumbai"}, timeout=15)
            item = next(it for it in r1.json()["items"] if it["product_id"] == pid)
            assert item["price"] == 150

            # Switch to Pune (clears cart), re-add
            cs.post(f"{API}/cart/set-city", json={"city": "Pune"}, timeout=15)
            r2 = cs.post(f"{API}/cart/add", json={"product_id": pid, "quantity": 1, "city": "Pune"}, timeout=15)
            item2 = next(it for it in r2.json()["items"] if it["product_id"] == pid)
            assert item2["price"] == 175
            cs.post(f"{API}/cart/clear", timeout=15)
        finally:
            aas.delete(f"{API}/products/mine/{pid}", timeout=15)
